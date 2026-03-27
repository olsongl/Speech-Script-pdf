#!/usr/bin/env python3
"""
Parse speech therapy log PDFs and reorganize into clean Excel files.

Per-file outputs (one set per PDF):
  out/<pdf_name>_cleaned.xlsx  — full log, color-coded by goal

Combined outputs (all PDFs merged, sorted by date):
  out/combined_cleaned.xlsx

Input PDFs go in:  <script_dir>/in/
Outputs are written to: <script_dir>/out/

Requirements: pip install pdfplumber pandas openpyxl
Usage: python parse_speech_logs.py         (uses in/ next to this script)
       python parse_speech_logs.py /path/  (overrides input folder)
"""

import re
import sys
from pathlib import Path

_SCRIPT_DIR  = Path(__file__).parent
INPUT_FOLDER = _SCRIPT_DIR / 'in'
OUTPUT_FOLDER = _SCRIPT_DIR / 'out'
from datetime import datetime, timedelta

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── session type code map ──────────────────────────────────────────────────────

SESSION_CODE_MAP = {
    'G':   'G - Groups',
    'I':   'I - Individual',
    'Ind': 'Ind - Indirect',
    'A':   'A - Student Absent',
    'B':   'B - School Closed',
    'C':   'C - Student Unavailable',
    'D':   'D - Clinician/Therapist Absent',
    'E':   'E - Clinician/Therapist Unavailable',
    'MT':  'MT - Music Therapy',
    'AT':  'AT - Art Therapy',
}


# ── goal color coding ──────────────────────────────────────────────────────────
# Each key must match what goal_category() returns.

GOAL_COLORS = {
    'Articulation':   'FFC7CE',   # soft red/pink
    'Expressive':     'C6EFCE',   # soft green
    'Receptive':      'BDD7EE',   # soft blue
    'Multiple Goals': 'FFE699',   # soft yellow
    'No Goal':        'D9D9D9',   # light gray
}

HEADER_FILL_HEX = '2F5496'  # dark blue for column headers


def goal_category(goal: str) -> str:
    """Return the color-coding category for a goal string."""
    g = goal.lower()
    active = [
        name for name, keyword in (
            ('Articulation', 'articulation'),
            ('Expressive',   'expressive'),
            ('Receptive',    'receptive'),
        )
        if keyword in g
    ]
    if not active:
        return 'No Goal'
    if len(active) > 1:
        return 'Multiple Goals'
    return active[0]


def monday_of_week(date: datetime) -> str:
    """Return MM/DD/YYYY of the Monday of the week containing *date*."""
    return (date - timedelta(days=date.weekday())).strftime('%m/%d/%Y')


# ── text cleanup ───────────────────────────────────────────────────────────────

_URL  = re.compile(r'https?://\S+\s*\d*/?\d*', re.IGNORECASE)
_JUNK = re.compile(
    r'^\s*[1-9]/[1-9]\s*$|'                    # standalone page numbers (e.g. "3/5") — NOT scores like "12/15"
    r'^\d+/\d+/\d+,\s+\d+:\d+.*$|'            # timestamp header (e.g. "6/17/25, 4:49 PM ...")
    r'^Date\s+Length.*$|'                      # column header rows
    r'^Absence\b.*$|'
    r'^Code\s*$|'
    r'^Provided and Related.*$|'
    r'^Session/.*$|'
    r'^\s*Initials/Supervisor.*$|'
    r'^\s*Electronic.*$|'
    r'^\s*(?:Signature\s*)+$|'
    r'^Service Nature\s+Number.*$|'
    r'^Speech/Language Therapy.*Weekly.*$|'
    r'^ESY:.*Weekly.*$',
    re.MULTILINE | re.IGNORECASE,
)


def clean_text(text: str) -> str:
    """Strip headers, footers, and table boilerplate from a page's text."""
    text = _URL.sub('', text)
    text = _JUNK.sub('', text)
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def extract_full_text(pdf_path: Path) -> tuple[str, str]:
    """Return (full_cleaned_text, first_page_raw_text)."""
    parts: list[str] = []
    first_raw = ''
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            raw = page.extract_text() or ''
            if i == 0:
                first_raw = raw
            parts.append(clean_text(raw))
    return '\n\n'.join(parts), first_raw


# ── header / metadata parsing ──────────────────────────────────────────────────

def extract_header_info(raw_first_page: str) -> dict:
    """Pull student name and provider name from page-1 header text."""
    name_m = re.search(
        r"Student'?s?\s+Name:\s*(.+?)\s+SASID:", raw_first_page, re.IGNORECASE
    )
    prov_m = re.search(
        r"Provider Name:\s*(.+?)\s+Position:", raw_first_page, re.IGNORECASE
    )
    return {
        'student_name':  name_m.group(1).strip() if name_m else '',
        'provider_name': prov_m.group(1).strip() if prov_m else '',
    }


# ── entry block parsing ────────────────────────────────────────────────────────

ENTRY_RE = re.compile(r'(?m)^(\d{2}/\d{2}/\d{4})\s+Start time:')


def split_into_blocks(text: str) -> list[tuple[str, str]]:
    """Return list of (date_str, block_text) for each session entry."""
    starts = [(m.start(), m.group(1)) for m in ENTRY_RE.finditer(text)]
    blocks = []
    for i, (pos, date) in enumerate(starts):
        end = starts[i + 1][0] if i + 1 < len(starts) else len(text)
        blocks.append((date, text[pos:end]))
    return blocks


def parse_block(date_str: str, block: str) -> dict:
    """Extract all fields from a single session entry block."""

    # ── times ─────────────────────────────────────────────────────────────────
    # In the PDF column layout, "pm" / "am" falls on the line after the digits.
    start_m = re.search(
        r'Start time:\s*(\d+:\d+).*?(\bam\b|\bpm\b)', block, re.DOTALL | re.IGNORECASE
    )
    end_m = re.search(
        r'End time:\s*(\d+:\d+).*?(\bam\b|\bpm\b)', block, re.DOTALL | re.IGNORECASE
    )
    start_t = (start_m.group(1) + start_m.group(2).lower()) if start_m else ''
    end_t   = (end_m.group(1)   + end_m.group(2).lower())   if end_m   else ''

    # ── session type ──────────────────────────────────────────────────────────
    first_line = block.split('\n')[0]
    code_m = re.search(r'Start time:\s*\d+:\d+\s+([A-Za-z]+)\s*-', first_line)
    session_type = SESSION_CODE_MAP.get(code_m.group(1), '') if code_m else ''

    # ── progress code & service location ──────────────────────────────────────
    prog_m = re.search(r'(?<!\w)([SPN])\s+(School|Home)\b', first_line, re.IGNORECASE)
    progress = prog_m.group(1).upper() if prog_m else ''
    location = prog_m.group(2).title()  if prog_m else ''

    # ── provider signature date ────────────────────────────────────────────────
    # The online system renders "Student Name: <name> <sig-date>" in the
    # Provider Electronic Signature column for each entry.
    sig_m = re.search(
        r'Student Name:.*?(\d{2}/\d{2}/\d{4})', block, re.DOTALL | re.IGNORECASE
    )
    provider_sig_date = sig_m.group(1) if sig_m else ''

    # ── goal ──────────────────────────────────────────────────────────────────
    goal_m = re.search(r'Goal:\s*(.*?)(?=Description:|$)', block, re.DOTALL | re.IGNORECASE)
    goal   = re.sub(r'\s+', ' ', goal_m.group(1)).strip() if goal_m else ''
    goal   = re.sub(r'\b\d{2}/\d{2}/\d{4}\b', '', goal).strip()
    goal   = re.sub(r'^End time:\s*\d+:\d+\s*', '', goal, flags=re.IGNORECASE).strip()
    goal   = re.sub(r'^(?:am|pm)\s*', '', goal, flags=re.IGNORECASE).strip()
    goal   = re.sub(r'Student Name:.*', '', goal, flags=re.IGNORECASE).strip()

    # ── actual description ────────────────────────────────────────────────────
    desc_m = re.search(r'Description:\s*(.*?)$', block, re.DOTALL | re.IGNORECASE)
    desc   = re.sub(r'\s+', ' ', desc_m.group(1)).strip() if desc_m else ''
    for stop_pat in (r'Date\s+Length of Session', r'Service Description:', r'Progress Code:'):
        desc = re.split(stop_pat, desc, flags=re.IGNORECASE)[0].strip()
    desc = re.sub(r'\b\d{2}/\d{2}/\d{4}\b', '', desc).strip()
    desc = re.sub(r'^\d+:\d+(?:\s*(?:am|pm))?(?:\s*-\s*\d+:\d+(?:\s*(?:am|pm))?)?\s*', '', desc, flags=re.IGNORECASE).strip()
    desc = re.sub(r'^(?:am|pm)\s*', '', desc, flags=re.IGNORECASE).strip()
    desc = re.sub(r'Student Name:.*', '', desc, flags=re.IGNORECASE).strip()

    return {
        'date':               datetime.strptime(date_str, '%m/%d/%Y'),
        'start_time':         start_t,
        'end_time':           end_t,
        'session_type':       session_type,
        'goal':               goal,
        'actual_description': desc,
        'progress_code':      progress,
        'service_location':   location,
        'provider_sig_date':  provider_sig_date,
    }


def parse_pdf(pdf_path: Path) -> tuple[list[dict], dict]:
    """Parse a speech-log PDF and return (entries, header_info)."""
    text, first_raw = extract_full_text(pdf_path)
    header          = extract_header_info(first_raw)
    blocks          = split_into_blocks(text)

    entries: list[dict] = []
    seen: set = set()
    for date_str, block in blocks:
        entry = parse_block(date_str, block)
        key   = (date_str, entry['start_time'], entry['session_type'][:20], entry['goal'][:40])
        if key in seen:
            continue
        seen.add(key)
        entries.append(entry)

    return entries, header


# ── build DataFrames ───────────────────────────────────────────────────────────

def build_dataframes(
    entries: list[dict], child_name: str
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (main_df, log_df)."""
    if not entries:
        return pd.DataFrame(), pd.DataFrame()

    df = pd.DataFrame(entries)
    df = df.sort_values('date').reset_index(drop=True)

    df['child_name']  = child_name
    df['day_of_week'] = df['date'].dt.strftime('%A')
    df['week_monday'] = df['date'].apply(monday_of_week)

    # "Start 2:00pm - End 2:30pm"
    def fmt_time(r):
        s, e = r['start_time'], r['end_time']
        if s and e:
            return f'Start {s} - End {e}'
        if s:
            return f'Start {s}'
        if e:
            return f'End {e}'
        return ''

    df['time'] = df.apply(fmt_time, axis=1)

    df['progress_code']      = df['progress_code'].apply(      lambda x: x if x else '(LEFT BLANK)')
    df['service_location']   = df['service_location'].apply(   lambda x: x if x else '(LEFT BLANK)')
    df['goal']               = df['goal'].apply(               lambda x: x if x else '(LEFT BLANK)')
    df['actual_description'] = df['actual_description'].apply( lambda x: x if x else '(LEFT BLANK)')
    df['goal_category']    = df['goal'].apply(goal_category)
    df['date_str']         = df['date'].dt.strftime('%m/%d/%Y')

    # ── main log ──────────────────────────────────────────────────────────────
    main_df = df[[
        'child_name', 'week_monday', 'date_str', 'provider_sig_date',
        'day_of_week',
        'time', 'session_type', 'goal', 'actual_description',
        'progress_code', 'service_location', 'goal_category',
    ]].copy()
    main_df.columns = [
        'Child Name', 'Week of (Monday)', 'Date of Service',
        'Provider Electronic Signature',
        'Day of Week',
        'Time', 'Session Type', 'Goal', 'Actual Description',
        'Progress Code', 'Service Location', 'Goal Category',
    ]

    # ── signature log ─────────────────────────────────────────────────────────
    log_df = df[['date_str', 'provider_sig_date']].copy()
    log_df.columns = ['Date of Service', 'Provider Electronic Signature']

    return main_df, log_df


# ── Excel formatting ───────────────────────────────────────────────────────────

# Fixed column widths for the main sheet (characters)
_MAIN_COL_WIDTHS = {
    'Child Name':         18,
    'Date of Service':    15,
    'Provider Electronic Signature': 28,
    'Day of Week':        13,
    'Week of (Monday)':   17,
    'Time':               30,
    'Session Type':       30,
    'Goal':               65,
    'Actual Description': 55,
    'Progress Code':      13,
    'Service Location':   15,
    'Goal Category':      17,
}


def style_main_sheet(ws, df: pd.DataFrame) -> None:
    """Apply header formatting and per-row goal color-coding."""
    header_fill = PatternFill(fill_type='solid', fgColor=HEADER_FILL_HEX)
    header_font = Font(bold=True, color='FFFFFF')

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        category  = row.get('Goal Category', 'No Goal')
        hex_color = GOAL_COLORS.get(category, 'FFFFFF')
        row_fill  = PatternFill(fill_type='solid', fgColor=hex_color)
        for cell in ws[row_idx]:
            cell.fill      = row_fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for i, col_name in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = _MAIN_COL_WIDTHS.get(col_name, 15)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'


def add_legend_sheet(wb) -> None:
    """Append a 'Legend' sheet explaining the goal color codes."""
    ws = wb.create_sheet('Legend')
    ws['A1'] = 'Goal Category'
    ws['B1'] = 'Color Code'
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for i, (cat, hex_color) in enumerate(GOAL_COLORS.items(), start=2):
        ws.cell(row=i, column=1, value=cat).fill  = PatternFill(fill_type='solid', fgColor=hex_color)
        ws.cell(row=i, column=2, value=f'#{hex_color}').fill = PatternFill(fill_type='solid', fgColor=hex_color)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12


def style_log_sheet(ws) -> None:
    """Apply simple formatting to the signature log sheet."""
    header_fill = PatternFill(fill_type='solid', fgColor=HEADER_FILL_HEX)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for col_cells in ws.columns:
        ws.column_dimensions[col_cells[0].column_letter].width = 28
    ws.freeze_panes = 'A2'


# ── save outputs ───────────────────────────────────────────────────────────────

def save_outputs(
    main_df: pd.DataFrame, log_df: pd.DataFrame, stem: Path
) -> tuple[Path, Path, Path]:
    """Write all output files and return (main_xlsx, log_xlsx, csv)."""

    # ── main cleaned Excel ────────────────────────────────────────────────────
    xlsx = stem.with_suffix('.xlsx')
    with pd.ExcelWriter(xlsx, engine='openpyxl') as writer:
        main_df.to_excel(writer, index=False, sheet_name='Speech Logs')
        style_main_sheet(writer.sheets['Speech Logs'], main_df)

    # Legend sheet must be added after ExcelWriter closes
    wb = load_workbook(xlsx)
    add_legend_sheet(wb)
    wb.save(xlsx)

    # ── signature log Excel ───────────────────────────────────────────────────
    log_xlsx = stem.parent / (stem.name.replace('_cleaned', '_log') + '.xlsx')
    with pd.ExcelWriter(log_xlsx, engine='openpyxl') as writer:
        log_df.to_excel(writer, index=False, sheet_name='Log')
        style_log_sheet(writer.sheets['Log'])

    # ── CSV (without internal Goal Category column) ───────────────────────────
    csv = stem.with_suffix('.csv')
    main_df.drop(columns=['Goal Category'], errors='ignore').to_csv(csv, index=False)

    return xlsx, log_xlsx, csv


# ── entry point ────────────────────────────────────────────────────────────────

def main():
    in_folder  = Path(sys.argv[1]) if len(sys.argv) > 1 else INPUT_FOLDER
    out_folder = Path(sys.argv[2]) if len(sys.argv) > 2 else OUTPUT_FOLDER

    if not in_folder.exists():
        print(f'Error: input folder not found: {in_folder}')
        sys.exit(1)

    out_folder.mkdir(parents=True, exist_ok=True)

    pdf_files = sorted(in_folder.glob('*.pdf'))
    if not pdf_files:
        print(f'Error: no PDF files found in {in_folder}')
        sys.exit(1)

    print(f'Found {len(pdf_files)} PDF(s) in {in_folder.resolve()}')
    print(f'Outputs -> {out_folder.resolve()}\n')

    all_main_dfs: list[pd.DataFrame] = []
    all_log_dfs:  list[pd.DataFrame] = []

    for pdf_path in pdf_files:
        print(f'Parsing {pdf_path.name} ...')
        entries, header = parse_pdf(pdf_path)
        print(f'  Extracted  : {len(entries)} entries')
        print(f'  Child      : {header["student_name"]}')
        print(f'  Provider   : {header["provider_name"]}')

        main_df, log_df = build_dataframes(entries, header['student_name'])

        stem = out_folder / (pdf_path.stem + '_cleaned')
        xlsx, log_xlsx, csv = save_outputs(main_df, log_df, stem)

        print(f'  Main Excel : {xlsx.name}')
        print(f'  Log Excel  : {log_xlsx.name}')
        print(f'  CSV        : {csv.name}')
        print()

        all_main_dfs.append(main_df)
        all_log_dfs.append(log_df)

    # ── combined outputs ───────────────────────────────────────────────────────
    if len(pdf_files) > 1:
        print('Building combined outputs ...')

        combined_main = (
            pd.concat(all_main_dfs, ignore_index=True)
            .assign(_sort=lambda d: pd.to_datetime(d['Date of Service'], format='%m/%d/%Y'))
            .sort_values('_sort')
            .drop(columns=['_sort'])
            .reset_index(drop=True)
        )
        combined_log = (
            pd.concat(all_log_dfs, ignore_index=True)
            .assign(_sort=lambda d: pd.to_datetime(d['Date of Service'], format='%m/%d/%Y'))
            .sort_values('_sort')
            .drop(columns=['_sort'])
            .reset_index(drop=True)
        )

        child_name = combined_main['Child Name'].iloc[0] if 'Child Name' in combined_main.columns and len(combined_main) else 'unknown'
        safe_name = re.sub(r'[^\w\s-]', '', child_name).strip().replace(' ', '_')
        today = datetime.today().strftime('%Y-%m-%d')
        combined_stem = out_folder / f'{safe_name}_combined_{today}'
        xlsx, log_xlsx, csv = save_outputs(combined_main, combined_log, combined_stem)

        print(f'  Combined Main Excel : {xlsx.name}')
        print(f'  Combined Log Excel  : {log_xlsx.name}')
        print(f'  Combined CSV        : {csv.name}')
        print(f'  Total entries       : {len(combined_main)}')
        print()


if __name__ == '__main__':
    main()
